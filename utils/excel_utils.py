from logging import Logger, getLogger

import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO

logger: Logger = getLogger(__name__)

async def create_excel_file(readings) -> BytesIO:
    """
    Создает Excel-файл с данными показаний счетчиков.

    Args:
        readings (list): Список объектов MeterReading.

    Returns:
        BytesIO: Excel-файл в формате BytesIO.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Показания счетчиков"

    # Заголовки
    headers = ["Квартира", "Тип счётчика", "Серийный номер" "Значение", "Дата подачи"]
    ws.append(headers)
    logger.info("Данные в ws: %s", ws)

    # Записываем данные
    for reading in readings:
        ws.append([
            reading["apartment_number"],
            reading["name"],
            reading["serial_number"],
            reading["value"],
            reading["reading_date"]
        ])

    logger.info("Данные в ws после: %s", ws)

    # Автоширина колонок
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        column_width = max(len(str(cell.value)) for cell in ws[column_letter]) + 2
        ws.column_dimensions[column_letter].width = column_width

    # Сохраняем файл в BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return excel_file
